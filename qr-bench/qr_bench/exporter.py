"""CSV and XLSX export for the accountant workflow. Pure formatting - no
DB access here; callers (main.py, via db.list_documents_for_export) must
hand in documents that already carry an `effective_fields` dict (original
parsed fields with any corrections applied on top - see
qr_bench.db.get_effective_fields).

Portuguese locale throughout: CSV is UTF-8 with a BOM (so Excel opens it
without a manual encoding prompt), semicolon-delimited, comma decimal
separator. XLSX uses native numeric/date cell types with PT-style number
formats rather than pre-formatted strings, so Excel's own locale handles
the actual on-screen rendering.
"""

import csv
import io
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from qr_bench import flags as flags_module
from qr_bench.export_presets import DEFAULT_PRESET, PRESETS
from qr_bench.validate import parse_amount

FIELD_KEY_PREFIX = "field_"
BASE_KEYS = ("I2", "I3", "I5", "I7")

FLAGGED_SHEET_HEADERS = ["Ficheiro", "Cliente", "Código", "Severidade", "Mensagem", "Causa", "Ação"]


def get_preset(preset_key: str) -> dict:
    return PRESETS.get(preset_key, PRESETS[DEFAULT_PRESET])


def _qr_key_for(column_key: str) -> str | None:
    if column_key.startswith(FIELD_KEY_PREFIX):
        return column_key[len(FIELD_KEY_PREFIX) :].upper()
    return None


def _bases_total(effective_fields: dict[str, str]) -> float | None:
    present = [b for k in BASE_KEYS if (b := parse_amount(effective_fields.get(k))) is not None]
    return sum(present) if present else None


def resolve_raw_value(doc: dict, column_key: str):
    effective_fields = doc.get("effective_fields") or {}
    if column_key == "bases_total":
        return _bases_total(effective_fields)
    if column_key in ("client_name", "supplier_name", "status"):
        return doc.get(column_key)
    qr_key = _qr_key_for(column_key)
    if qr_key:
        return effective_fields.get(qr_key)
    return doc.get(column_key)


def format_number_pt(value: float) -> str:
    """1234.5 -> '1234,50' - comma decimal, no thousands separator (kept
    plain/unambiguous for accounting-software CSV import)."""
    return f"{value:.2f}".replace(".", ",")


def format_date_pt(yyyymmdd: str | None) -> str:
    if not yyyymmdd:
        return ""
    try:
        return datetime.strptime(yyyymmdd, "%Y%m%d").strftime("%d/%m/%Y")
    except ValueError:
        return yyyymmdd


def parse_date_value(yyyymmdd: str | None) -> date | None:
    if not yyyymmdd:
        return None
    try:
        return datetime.strptime(yyyymmdd, "%Y%m%d").date()
    except ValueError:
        return None


def _formatted_cell_csv(value, column_type: str) -> str:
    if value is None or value == "":
        return ""
    if column_type == "currency":
        parsed = value if isinstance(value, (int, float)) else parse_amount(str(value))
        return format_number_pt(parsed) if parsed is not None else ""
    if column_type == "date":
        return format_date_pt(str(value))
    return str(value)


def build_csv(documents: list[dict], preset_key: str = DEFAULT_PRESET) -> bytes:
    preset = get_preset(preset_key)
    columns = preset["columns"]

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow([c["header_pt"] for c in columns])
    for doc in documents:
        writer.writerow(
            _formatted_cell_csv(resolve_raw_value(doc, col["key"]), col["type"]) for col in columns
        )

    return ("﻿" + buffer.getvalue()).encode("utf-8")


def _autosize(ws, col_count: int, min_widths: list[int]) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=col_count):
        for cell in row:
            length = len(str(cell.value)) if cell.value is not None else 0
            idx = cell.column - 1
            if length > min_widths[idx]:
                min_widths[idx] = length
    for i, width in enumerate(min_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 80)


def _write_documents_sheet(ws, documents: list[dict], columns: list[dict]) -> None:
    headers = [c["header_pt"] for c in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for doc in documents:
        row_values = []
        for col in columns:
            raw = resolve_raw_value(doc, col["key"])
            if col["type"] == "currency":
                parsed = raw if isinstance(raw, (int, float)) else parse_amount(str(raw)) if raw else None
                row_values.append(parsed)
            elif col["type"] == "date":
                row_values.append(parse_date_value(str(raw)) if raw else None)
            else:
                row_values.append(raw if raw is not None else "")
        ws.append(row_values)

        row_idx = ws.max_row
        for i, col in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=i)
            if col["type"] == "currency":
                cell.number_format = '#,##0.00 €'
            elif col["type"] == "date":
                cell.number_format = "DD/MM/YYYY"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws, len(columns), [len(h) for h in headers])


def _write_flagged_sheet(ws, documents: list[dict]) -> None:
    ws.append(FLAGGED_SHEET_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for doc in documents:
        for flag in doc.get("validation_flags", []):
            registry_entry = flags_module.FLAGS.get(flag["code"], {})
            ws.append(
                [
                    doc.get("filename", ""),
                    doc.get("client_name") or "",
                    flag["code"],
                    flag.get("severity", ""),
                    flag.get("message", ""),
                    registry_entry.get("cause_pt", ""),
                    registry_entry.get("action_pt", ""),
                ]
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws, len(FLAGGED_SHEET_HEADERS), [len(h) for h in FLAGGED_SHEET_HEADERS])


def build_xlsx(documents: list[dict], preset_key: str = DEFAULT_PRESET) -> bytes:
    preset = get_preset(preset_key)
    columns = preset["columns"]

    wb = Workbook()
    documents_sheet = wb.active
    documents_sheet.title = "Documentos"
    _write_documents_sheet(documents_sheet, documents, columns)

    flagged_sheet = wb.create_sheet("Documentos Sinalizados")
    _write_flagged_sheet(flagged_sheet, documents)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_filename(scope_label: str, period_label: str, extension: str) -> str:
    return f"clientes_{scope_label}_{period_label}.{extension}"


def month_bounds(year: int, month: int) -> tuple[str, str]:
    """First/last day of the given month as YYYYMMDD strings."""
    start = date(year, month, 1)
    next_month = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    end = next_month - timedelta(days=1)
    return start.strftime("%Y%m%d"), end.strftime("%Y%m%d")
