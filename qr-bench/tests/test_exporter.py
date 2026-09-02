"""Tests for CSV/XLSX export: Portuguese-locale formatting and round-tripping
(export a document, read the file back, confirm the values match)."""

import csv
import datetime
import io
import random
import sys
from pathlib import Path

from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))
from generate_samples import build_qr_payload, render_invoice_canvas  # noqa: E402

from qr_bench import db, exporter, pipeline  # noqa: E402


def _process(rng, **kwargs):
    payload = build_qr_payload(rng, **kwargs)
    image = render_invoice_canvas(payload.raw, rng)
    doc = pipeline.process_image(image, "test.png", "test", page_number=None)
    return doc, payload


def _export_all(include_unvalidated: bool = False):
    return db.list_documents_for_export(
        client_id=None,
        unattributed_only=False,
        date_from=None,
        date_to=None,
        include_unvalidated=include_unvalidated,
    )


def test_format_number_pt_uses_comma_decimal_no_thousands_separator():
    assert exporter.format_number_pt(1234.5) == "1234,50"
    assert exporter.format_number_pt(0) == "0,00"
    assert exporter.format_number_pt(9.999) == "10,00"


def test_format_date_pt_is_ddmmyyyy():
    assert exporter.format_date_pt("20260315") == "15/03/2026"


def test_filename_matches_the_spec_pattern():
    assert exporter.build_filename("509442083", "2026-08", "xlsx") == "clientes_509442083_2026-08.xlsx"
    assert exporter.build_filename("todos", "2026-08", "csv") == "clientes_todos_2026-08.csv"


def test_csv_export_has_bom_semicolon_delimiter_and_comma_decimals(test_db):
    rng = random.Random(1)
    doc, _ = _process(rng)
    db.set_document_status(doc["id"], "validado")

    csv_bytes = exporter.build_csv(_export_all())

    assert csv_bytes.startswith(b"\xef\xbb\xbf")
    text = csv_bytes.decode("utf-8-sig")

    rows = list(csv.reader(io.StringIO(text), delimiter=";"))
    assert len(rows) == 2  # header + one document
    header, data_row = rows
    assert "Total" in header

    total_value = data_row[header.index("Total")]
    assert "," in total_value
    assert "." not in total_value


def test_csv_round_trips_the_corrected_total_not_the_original(test_db):
    rng = random.Random(2)
    doc, _ = _process(rng)
    db.record_corrections(doc["id"], {"O": "555.25"}, "Rui")  # also -> validado

    csv_bytes = exporter.build_csv(_export_all())
    text = csv_bytes.decode("utf-8-sig")
    header, data_row = list(csv.reader(io.StringIO(text), delimiter=";"))

    assert data_row[header.index("Total")] == "555,25"


def test_csv_default_scope_excludes_unvalidated_documents(test_db):
    rng = random.Random(3)
    doc, _ = _process(rng)  # status extraido, never validated

    docs_default = _export_all(include_unvalidated=False)
    assert doc["id"] not in [d["id"] for d in docs_default]

    docs_all = _export_all(include_unvalidated=True)
    assert doc["id"] in [d["id"] for d in docs_all]


def test_xlsx_export_uses_real_dates_and_numeric_cells(test_db):
    rng = random.Random(4)
    doc, _ = _process(rng)
    db.set_document_status(doc["id"], "validado")

    xlsx_bytes = exporter.build_xlsx(_export_all())
    wb = load_workbook(io.BytesIO(xlsx_bytes))

    assert wb.sheetnames == ["Documentos", "Documentos Sinalizados"]
    ws = wb["Documentos"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None

    header = [cell.value for cell in ws[1]]
    date_col = header.index("Data")
    total_col = header.index("Total")
    data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    assert isinstance(data_row[date_col], (datetime.date, datetime.datetime))
    assert isinstance(data_row[total_col], (int, float))


def test_xlsx_round_trips_the_corrected_total(test_db):
    rng = random.Random(5)
    doc, _ = _process(rng)
    db.record_corrections(doc["id"], {"O": "321.00"}, "Ana")

    xlsx_bytes = exporter.build_xlsx(_export_all())
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Documentos"]
    header = [cell.value for cell in ws[1]]
    total_col = header.index("Total")
    data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    assert data_row[total_col] == 321.0


def test_xlsx_flagged_sheet_lists_flag_codes_and_registry_text(test_db):
    rng = random.Random(6)
    doc, _ = _process(rng, break_totals=True)
    db.set_document_status(doc["id"], "validado")

    xlsx_bytes = exporter.build_xlsx(_export_all())
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    flagged_ws = wb["Documentos Sinalizados"]

    header = [cell.value for cell in flagged_ws[1]]
    rows = list(flagged_ws.iter_rows(min_row=2, values_only=True))
    codes = [row[header.index("Código")] for row in rows]
    assert "TOTAIS_INCONSISTENTES" in codes

    cause_col = header.index("Causa")
    action_col = header.index("Ação")
    flagged_row = rows[codes.index("TOTAIS_INCONSISTENTES")]
    assert flagged_row[cause_col]  # registry text present, not blank
    assert flagged_row[action_col]


def test_export_via_api_marks_documents_exported_with_timestamp(test_db):
    import asyncio

    from qr_bench import main

    rng = random.Random(7)
    doc, _ = _process(rng)
    db.set_document_status(doc["id"], "validado")

    payload = main.ExportPayload(
        format="csv",
        scope="all",
        client_id=None,
        period=main.ExportPeriod(mode="range", start="2000-01-01", end="2099-12-31"),
        status_mode="validado_only",
        preset="generico",
    )
    response = asyncio.run(main.export_documents(payload))
    assert response.status_code == 200
    assert response.body.startswith(b"\xef\xbb\xbf")

    refreshed = db.get_document(doc["id"])
    assert refreshed["status"] == "exportado"
    assert refreshed["exported_at"] is not None


def test_reexport_via_api_is_allowed(test_db):
    import asyncio

    from qr_bench import main

    rng = random.Random(8)
    doc, _ = _process(rng)
    db.set_document_status(doc["id"], "validado")

    payload = main.ExportPayload(
        format="csv",
        scope="all",
        client_id=None,
        period=main.ExportPeriod(mode="range", start="2000-01-01", end="2099-12-31"),
        status_mode="include_unvalidated",
        preset="generico",
    )
    asyncio.run(main.export_documents(payload))
    first_export = db.get_document(doc["id"])["exported_at"]

    second_response = asyncio.run(main.export_documents(payload))
    assert second_response.status_code == 200
    second_export = db.get_document(doc["id"])["exported_at"]
    assert second_export >= first_export
